#!/usr/bin/env python3
""" Streamlit web app version of the original Google Colab workflow.
Logic and formulas are preserved.
"""

import io
import os
import hashlib
from typing import Dict, Optional

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import streamlit as st
from bisect import bisect_left
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table

# =========================
# HELPER FUNCTIONS
# =========================

def find_exact_column(columns, name):
    """Find column by exact normalized name (prevents kV vs kV 2 confusion)."""
    target = name.lower()
    for col in columns:
        if str(col).strip().lower() == target:
            return col
    return None


def safe_col(df, col):
    """Convert safely to numeric and replace NaN with 0."""
    if col is None:
        return pd.Series([0] * len(df), index=df.index)
    return pd.to_numeric(df[col], errors="coerce").fillna(0)


def has_values(df, col):
    if col is None:
        return False
    return df[col].notna().any()


def print_status(text, color):
    if color == "#1f6feb":
        st.info(text)
    elif color == "orange":
        st.warning(text)
    elif color == "red":
        st.error(text)
    else:
        st.success(text)


def print_info(text):
    print_status(text, "#1f6feb")


def print_warning(text):
    print_status(text, "orange")


def print_error(text):
    print_status(text, "red")


def print_download_message():
    st.success("PDF wurde erfolgreich erstellt.")


def is_invalid(series):
    s = series.dropna()
    total_len = len(s)

    if total_len == 0:
        return True

    zero_count = (s == 0).sum()

    return zero_count == total_len or zero_count >= (total_len / 3)


def calculate_weeks(von, bis):
    von = pd.to_datetime(von)
    bis = pd.to_datetime(bis)
    return ((bis - von).days + 1) / 7


def require(condition, message):
    """Raise a user-readable error if condition fails."""
    if not condition:
        raise ValueError(message)


# =========================
# CONVERSION TABLE
# =========================

conversion_table = {
    0: [1, 1.4, 1.7, 2.1, 2.5, 3.2, 4.2, 5.8, 9],
    0.2: [1, 1.7, 2.5, 3.6, 5.4, 10, 26, 79, 270],
    0.5: [1, 1.8, 2.8, 4.2, 7.1, 17, 70, 440, 3800],
    1: [1, 1.8, 2.8, 4.6, 8.8, 31, 290, 7400, 10000],
    1.5: [1, 1.8, 2.9, 5, 11, 58, 1400, 10000, None],
    2: [1, 1.8, 3.1, 5.5, 14, 110, 7000, None, None],
    2.5: [1, 1.8, 3.3, 6.2, 18, 230, 10000, None, None],
    3: [1, 1.9, 3.5, 7.2, 24, 470, None, None, None],
    3.5: [1, 1.9, 3.8, 8.3, 33, 990, None, None, None],
}

kv_cols = [150, 125, 110, 100, 90, 80, 70, 60, 50]


def _table_factor_for_kv(mm: float, kv_value: float) -> Optional[float]:
    """Return factor for a given mm Pb and kV using linear interpolation when needed."""
    row = conversion_table.get(mm)
    if row is None:
        return None

    # Build valid (kV, factor) points for this mm Pb row.
    points = [(kv, fac) for kv, fac in zip(kv_cols, row) if fac is not None]
    if not points:
        return None

    # Sort ascending by kV for interpolation.
    points.sort(key=lambda x: x[0])
    kvs = [p[0] for p in points]
    facs = [p[1] for p in points]

    # Exact match.
    if kv_value in kvs:
        return facs[kvs.index(kv_value)]

    # Outside table range -> use nearest available kV in the row.
    if kv_value <= kvs[0]:
        return facs[0]
    if kv_value >= kvs[-1]:
        return facs[-1]

    # Linear interpolation between surrounding columns.
    i = bisect_left(kvs, kv_value)
    kv_lo, kv_hi = kvs[i - 1], kvs[i]
    f_lo, f_hi = facs[i - 1], facs[i]
    ratio = (kv_value - kv_lo) / (kv_hi - kv_lo)
    return f_lo + ratio * (f_hi - f_lo)


def get_factor(kv_ref: float, kv_override: float, mm: float):
    """
    Required rule:
    factor = factor(kV-ref) / factor(kV-override)
    where each factor comes from conversion table (with interpolation if needed).
    """
    factor_avg = _table_factor_for_kv(mm, kv_ref)
    factor_user = _table_factor_for_kv(mm, kv_override)

    if factor_avg is None or factor_user is None or factor_user <= 0:
        return None, factor_avg, factor_user

    return factor_avg / factor_user, factor_avg, factor_user


# =========================
# METADATA HELPERS
# =========================

def get_best_value_from_columns(df, label):
    label_lower = label.lower().strip()
    matching_cols = [c for c in df.columns if label_lower in str(c).lower()]

    found_values = []
    for col in matching_cols:
        vals = df[col].dropna().astype(str).tolist()
        clean_vals = [
            v.strip()
            for v in vals
            if v.lower() not in ["nan", "none", "", "0", "0.0"]
        ]
        if clean_vals:
            found_values.append(clean_vals[0])

    if found_values:
        return min(list(set(found_values)), key=len)
    return "Nicht gefunden"


def render_available_parameters_table(df_data: pd.DataFrame, modality: str):
    if "CT" in modality:
        param_list = [
            ("Röhrenstrom Durchschnitt (mA)", "Röhrenstrom Durchschnitt (mA)"),
            ("Röhrenstrom Durchschnitt 2 (mA)", "Röhrenstrom Durchschnitt 2 (mA)"),
            ("Bestrahlungszeit (s)", "Bestrahlungszeit (s)"),
        ]
    else:
        param_list = [
            ("Bestrahlungszeit (ms)", "Bestrahlungszeit (ms)"),
            ("Bestrahlungsdauer (s)", "Bestrahlungsdauer (s)"),
            ("Anzahl der Impulse", "Anzahl der Impulse"),
            ("Röntgenröhrenstrom (mA)", "Röntgenröhrenstrom (mA)"),
            ("Impulsbreite (ms)", "Impulsbreite (ms)"),
            ("Bestrahlungsintensität (µAs)", "Bestrahlungsintensität (µAs)"),
        ]

    rows = []
    for name, col_name in param_list:
        col_label = find_exact_column(df_data.columns, col_name)
        status = "✔" if col_label is not None and not is_invalid(df_data[col_label]) else "✘"
        rows.append({"Parameter": name, "Verfügbar": status})

    st.subheader("Verfügbare Parameter")
    st.table(pd.DataFrame(rows))


# =========================
# CHART
# =========================

def create_chart(value, allowed, out_dir):
    percent = (value / allowed) * 100 if allowed > 0 else 0

    plt.figure(figsize=(6, 4))
    bars = plt.bar(["Berechnet", "Bewilligt"], [value, allowed])

    plt.text(
        0.5,
        1.08,
        f"Auslastung: {percent:.1f}%",
        ha="center",
        va="bottom",
        transform=plt.gca().transAxes,
        fontsize=12,
    )

    for bar in bars:
        height = bar.get_height()
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{height:.1f}",
            ha="center",
            va="bottom",
            fontsize=10,
        )

    path = os.path.join(out_dir, "chart.png")
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


# =========================
# PDF
# =========================

def create_pdf(
    modality,
    weekly,
    allowed,
    final_val,
    von,
    bis,
    info_dict,
    kv_text,
    conversion,
    out_pdf,
    df_data,
    kv_ref,
    kv_input
):
    out_dir = os.path.dirname(out_pdf) or "."
    file_name = os.path.basename(out_pdf)

    doc = SimpleDocTemplate(
        filename=file_name,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = []

    von_dt = pd.to_datetime(von)
    bis_dt = pd.to_datetime(bis)
    weeks = calculate_weeks(von_dt, bis_dt)

    elements.append(Paragraph("Betriebsumfang-Auswertung", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Modalität: {modality}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"Zeitraum: {von_dt.strftime('%Y-%m-%d')} bis {bis_dt.strftime('%Y-%m-%d')}",
            styles["Normal"],
        )
    )
    elements.append(Paragraph(f"Anzahl der Wochen: {weeks:.2f}", styles["Normal"]))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph("Geräteinformationen:", styles["Heading2"]))
    elements.append(
        Paragraph(
            (
                f"Gerät: {info_dict['Gerät']}, "
                f"Standort: {info_dict['Standort']}, "
                f"Hersteller: {info_dict['Hersteller']}, "
                f"Modell: {info_dict['Modell']}"
            ),
            styles["Normal"],
        )
    )

    if kv_text:
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(kv_text, styles["Normal"]))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Verfügbare Parameter:", styles["Heading2"]))

    param_table = [["", ""]]
    if "CT" in modality:
        param_list = [
            ("Röhrenstrom Durchschnitt (mA)", "Röhrenstrom Durchschnitt (mA)"),
            ("Röhrenstrom Durchschnitt 2 (mA)", "Röhrenstrom Durchschnitt 2 (mA)"),
            ("Bestrahlungszeit (s)", "Bestrahlungszeit (s)"),
        ]
    else:
        param_list = [
            ("Bestrahlungszeit (ms)", "Bestrahlungszeit (ms)"),
            ("Bestrahlungsdauer (s)", "Bestrahlungsdauer (s)"),
            ("Anzahl der Impulse", "Anzahl der Impulse"),
            ("Röntgenröhrenstrom (mA)", "Röntgenröhrenstrom (mA)"),
            ("Impulsbreite (ms)", "Impulsbreite (ms)"),
            ("Bestrahlungsintensität (µAs)", "Bestrahlungsintensität (µAs)"),
        ]

    for name, col_name in param_list:
        col_label = find_exact_column(df_data.columns, col_name)
        status = "✔" if col_label is not None and not is_invalid(df_data[col_label]) else "✘"
        param_table.append([name, status])

    elements.append(Table(param_table))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Ergebnis:", styles["Heading2"]))

    status = "INNERHALB" if weekly <= allowed else "ÜBERSCHRITTEN"
    color = "green" if status == "INNERHALB" else "red"

    elements.append(Paragraph(f"Berechneter Betriebsumfang: {weekly:.2f}", styles["Normal"]))
    elements.append(Paragraph(f"Bewilligter Betriebsumfang: {allowed:.2f}", styles["Normal"]))
    elements.append(
        Paragraph(f"<font color='{color}'><b>Status: {status}</b></font>", styles["Normal"])
    )

    if conversion:
        mm, factor, new_val, new_status, _ = conversion
        elements.append(Spacer(1, 10))
        elements.append(
            Paragraph(
                f"Nennspannung: {kv_input:.2f}, Durchschnittlicher kV: {kv_ref:.2f}",
                styles["Normal"],
            )
        )
        elements.append(
            Paragraph(
                f"Bleigleichwert: {mm} mm Pb, Umrechnungsfaktor: {factor:.2f}",
                styles["Normal"],
            )
        )
        elements.append(
            Paragraph(
                f"Berechneter Betriebsumfang nach ÖNORM S 5212: {new_val:.2f}",
                styles["Normal"],
            )
        )
        color2 = "green" if new_status == "INNERHALB" else "red"
        elements.append(
            Paragraph(
                f"<font color='{color2}'><b>Status nach Umrechnung: {new_status}</b></font>",
                styles["Normal"],
            )
        )

    elements.append(Spacer(1, 12))
    elements.append(Image(create_chart(final_val, allowed, out_dir), width=350, height=220))

    doc.build(elements)


# =========================
# LOAD FILE (OPTIMIZED IN-PLACE)
# =========================

@st.cache_data(show_spinner=False)
def load_excel_data(excel_bytes: bytes):
    progress_bar = st.progress(0, text="0% - Initialisiere Datei...")

    # Fast header scan: stream only the first sheet's first 25 rows.
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    scan_rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=25, values_only=True)]
    wb.close()
    raw_df = pd.DataFrame(scan_rows)
    progress_bar.progress(15, text="15% - Datei geladen")

    progress_bar.progress(35, text="35% - Suche Tabellenkopf...")
    header_row_index = None
    anchor_keywords = ["kv", "ma", "zeit", "ms", "intensität", "puls", "µas"]

    for idx, row in raw_df.iterrows():
        # FIX: Ensure every single item is a string and handle NaN explicitly
        clean_row_values = [str(val).lower() if val is not None else "" for val in row]
        row_string = " ".join(clean_row_values)
        matches = [key for key in anchor_keywords if key in row_string]
        if len(matches) >= 2:
            header_row_index = idx
            break

    if header_row_index is None:
        header_row_index = 14
        progress_bar.progress(50)
        print_warning(
            f"Warnung: Header nicht automatisch gefunden. Nutze Zeile {header_row_index}."
        )

    progress_bar.progress(65, text="65% - Tabelle wird eingelesen...")
    # Prefer calamine for speed if available; fallback to openpyxl.
    try:
        df_data = pd.read_excel(
            io.BytesIO(excel_bytes),
            header=header_row_index,
            engine="calamine",
        )
    except Exception:
        df_data = pd.read_excel(
            io.BytesIO(excel_bytes),
            header=header_row_index,
            engine="openpyxl",
        )
    df_data.columns = df_data.columns.astype(str).str.strip()

    progress_bar.progress(100, text="100% - Fertig!")
    progress_bar.empty()

    return raw_df, df_data


def calculate(
    raw_df: pd.DataFrame,
    df_data: pd.DataFrame,
    allowed: float,
    mm_pb: Optional[float] = None,
    kv_override: Optional[float] = None,
    out_pdf: str = "Betriebsumfang_Report.pdf",
    write_pdf: bool = True,
):
    require(allowed > 0, "Bitte einen gültigen Wert für mAmin/W > 0 eingeben.")

    von = raw_df.iloc[2, 2]
    bis = raw_df.iloc[3, 2]
    require(pd.notna(von) and pd.notna(bis), "Zeitraum (von/bis) konnte nicht gelesen werden.")

    weeks = calculate_weeks(von, bis)
    require(weeks > 0, "Ungültiger Zeitraum: 'bis' liegt vor 'von'.")

    modality = str(raw_df.iloc[4, 2])
    require(modality and modality.lower() != "nan", "Modalität konnte nicht gelesen werden.")

    df = df_data

    info_dict = {
        "Gerät": get_best_value_from_columns(df, "Gerät"),
        "Standort": get_best_value_from_columns(df, "Standort"),
        "Hersteller": get_best_value_from_columns(df, "Hersteller"),
        "Modell": get_best_value_from_columns(df, "Modell"),
    }

    # Faster: resolve needed columns once.
    wanted = {
        "mA_avg": "Röhrenstrom Durchschnitt (mA)",
        "mA2": "Röhrenstrom Durchschnitt 2 (mA)",
        "time_CT": "Bestrahlungszeit (s)",
        "time": "Bestrahlungszeit (ms)",
        "mA": "Röntgenröhrenstrom (mA)",
        "w": "Impulsbreite (ms)",
        "n": "Anzahl der Impulse",
        "int": "Bestrahlungsintensität (µAs)",
        "kv1": "kVp (kV)",
        "kv2": "kVp 2 (kV)",
        "level": "Akquisitionsebene",
    }
    cols = {k: find_exact_column(df.columns, v) for k, v in wanted.items()}

    if any(m in modality for m in ["CT", "NM/CT", "PT/CT"]):
        current_sum = safe_col(df, cols["mA_avg"]).clip(lower=0) + safe_col(df, cols["mA2"]).clip(lower=0)
        time_vals = safe_col(df, cols["time_CT"]).clip(lower=0)
        total = (current_sum * time_vals).sum()
        avg_time = time_vals.mean()

        if avg_time > 100:
            weekly = total / (60 * weeks * 1000)
        else:
            weekly = total / (60 * weeks)

    elif any(m in modality for m in ["CR", "DX", "MG", "RF", "XA"]):
        mA_vals = safe_col(df, cols["mA"])
        time_vals = safe_col(df, cols["time"])
        w_vals = safe_col(df, cols["w"])
        n_vals = safe_col(df, cols["n"])
        int_vals = safe_col(df, cols["int"])

        if not (is_invalid(mA_vals) or is_invalid(time_vals)):
            total = (mA_vals.clip(lower=0) * time_vals.clip(lower=0)).sum()
        elif not (is_invalid(w_vals) or is_invalid(n_vals) or is_invalid(mA_vals)):
            total = (w_vals.clip(lower=0) * n_vals.clip(lower=0) * mA_vals.clip(lower=0)).sum()
        elif not is_invalid(int_vals):
            total = int_vals.clip(lower=0).sum()
        else:
            raise ValueError(
                "Die Parameter für die Berechnung des Betriebsumfangs sind in der hochgeladenen Excel-Datei nicht verfügbar."
            )

        weekly = total / (1000 * 60 * weeks)
    else:
        raise ValueError(
            f"Unbekannte Modalität '{modality}'. Unterstützt: CT/NM/CT/PT/CT oder CR/DX/MG/RF/XA."
        )

    kv_text_lines = []
    kv_ref = 0

    if "CT" in modality:
        kv1_vals = pd.to_numeric(df[cols["kv1"]], errors="coerce").dropna() if cols["kv1"] else pd.Series([])
        kv1_vals = kv1_vals[kv1_vals > 0]

        kv2_vals = pd.to_numeric(df[cols["kv2"]], errors="coerce").dropna() if cols["kv2"] else pd.Series([])
        kv2_vals = kv2_vals[kv2_vals > 0]

        count1 = (kv1_vals != 0).sum()
        count2 = (kv2_vals != 0).sum()

        kv_ref = kv2_vals.mean() if count2 > count1 and count2 > 0 else (kv1_vals.mean() if count1 > 0 else 0)

        if len(kv1_vals) > 0:
            kv_text_lines.append(f"Durchschnittlicher kVp: {kv1_vals.mean():.2f}")
        if len(kv2_vals) > 0:
            kv_text_lines.append(f"Durchschnittlicher kVp 2: {kv2_vals.mean():.2f}")
    else:
        if cols["level"] and has_values(df, cols["level"]):
            groups = df.groupby(cols["level"])
            counts: Dict[str, int] = {}
            for name, group in groups:
                kv_vals = pd.to_numeric(group[cols["kv1"]], errors="coerce").dropna() if cols["kv1"] else pd.Series([])
                kv_vals = kv_vals[kv_vals > 0]
                if len(kv_vals) > 0:
                    kv_text_lines.append(f"kV {name}: {kv_vals.mean():.2f}")
                    counts[name] = len(kv_vals)

            if counts and cols["kv1"]:
                main_group = max(counts, key=counts.get)
                ref_vals = pd.to_numeric(df[df[cols["level"]] == main_group][cols["kv1"]], errors="coerce").dropna()
                ref_vals = ref_vals[ref_vals > 0]
                kv_ref = ref_vals.mean() if len(ref_vals) > 0 else 0
        elif cols["kv1"]:
            kv_vals = pd.to_numeric(df[cols["kv1"]], errors="coerce").dropna()
            kv_vals = kv_vals[kv_vals > 0]
            if len(kv_vals) > 0:
                kv_ref = kv_vals.mean()
                kv_text_lines.append(f"Durchschnittlicher kVp: {kv_ref:.2f}")

    kv_text = "<br/>".join(kv_text_lines)
    status = "INNERHALB" if weekly <= allowed else "ÜBERSCHRITTEN"

    final_val = weekly

    conversion = None

    if status == "ÜBERSCHRITTEN" and mm_pb is not None:
        factor, factor_avg, factor_user = get_factor(kv_ref, kv_override, mm_pb)
        require(kv_ref > 0, "Kein gültiger durchschnittlicher kV-Wert in den Daten gefunden.")
        require(kv_override is not None and kv_override > 0, "Ungültiger eingegebener kV-Wert.")

        new_val = weekly / factor
        new_status = "INNERHALB" if new_val <= allowed else "ÜBERSCHRITTEN"
        conversion = (mm_pb, factor, new_val, new_status, kv_text)
        final_val = new_val

    if write_pdf:
        create_pdf(
            modality,
            weekly,
            allowed,
            final_val,
            von,
            bis,
            info_dict,
            kv_text,
            conversion,
            out_pdf,
            df_data,
            kv_ref,
            kv_override,

        )

    return {
        "weekly": weekly,
        "allowed": allowed,
        "status": status,
        "modality": modality,
        "von": von,
        "bis": bis,
        "info_dict": info_dict,
        "kv_text": kv_text,
        "conversion": conversion,
        "out_pdf": out_pdf,
        "final_val": final_val,
    }


def app():
    st.set_page_config(page_title="Betriebsumfang-Auswertung", layout="centered")
    st.title("Betriebsumfang-Auswertung")
    print_info("Schritt 1: Laden Sie Ihre Excel-Datei hoch und geben Sie den Betriebsumfang sowie die Nennspannung ein.")
    print_info("Schritt 2: Klicken Sie auf „Berechnen“. Ihr PDF wird automatisch erstellt und heruntergeladen.")
    print_info("Hinweis: Falls eine Umrechnung erforderlich ist, geben Sie im nächsten Schritt den Bleigleichwert an und klicken Sie auf „Umrechnen“.")

    if "mm_pb_input" not in st.session_state:
        st.session_state.mm_pb_input = 0.0
    if "calc_result" not in st.session_state:
        st.session_state.calc_result = None
    if "calc_done" not in st.session_state:
        st.session_state.calc_done = False

    def reset_for_next_entry():
        st.session_state.mm_pb_input = 0.0
        st.session_state.calc_result = None
        st.session_state.calc_done = False

    uploaded_file = st.file_uploader("", type=["xlsx", "xlsm", "xls"])
    st.caption("Datei hierher ziehen oder auf „Browse files“ klicken.")
    if uploaded_file is None:
        return

    input_col1, mid_col, input_col2 = st.columns([2, 1, 2])
    with input_col1:
        allowed = st.number_input("mAmin/W:", min_value=0.0, value=0.0, step=0.1)
    with mid_col:
        st.markdown("<div style='padding-top:2rem;text-align:center;'>bei</div>", unsafe_allow_html=True)
    with input_col2:
        kv_input = st.number_input("kV:", min_value=0.0, value=0.0, step=1.0)

    excel_bytes = uploaded_file.getvalue()
    file_hash = hashlib.md5(excel_bytes).hexdigest()

    # Parse once per uploaded file and keep in session to avoid expensive reloads on reruns.
    if st.session_state.get("loaded_file_hash") != file_hash:
        try:
            raw_df, df_data = load_excel_data(excel_bytes)
            st.session_state.loaded_file_hash = file_hash
            st.session_state.loaded_raw_df = raw_df
            st.session_state.loaded_df_data = df_data
        except Exception as err:
            print_error(f"Datei konnte nicht gelesen werden: {err}")
            return
    else:
        raw_df = st.session_state.loaded_raw_df
        df_data = st.session_state.loaded_df_data

    run_calc = st.button("Berechnen", type="primary")

    if run_calc:
        try:
            require(allowed > 0, "Bitte mAmin/Woche eingeben (Wert > 0).")
            require(kv_input > 0, "Bitte kV eingeben (Wert > 0).")
            result = calculate(
                raw_df,
                df_data,
                allowed=allowed,
                mm_pb=None,
                kv_override=kv_input,
                out_pdf="Betriebsumfang_Report.pdf",
                write_pdf=False,
            )
            st.session_state.calc_result = result
            st.session_state.calc_done = True
        except Exception as err:
            st.session_state.calc_done = False
            st.session_state.calc_result = None
            print_error(f"Fehler: {err}")
            if "Die Parameter für die Berechnung des Betriebsumfangs sind in der hochgeladenen Excel-Datei nicht verfügbar." in str(err):
                modality_hint = str(raw_df.iloc[4, 2]) if len(raw_df.index) > 4 else ""
                render_available_parameters_table(df_data, modality_hint)

    if not st.session_state.calc_done or not st.session_state.calc_result:
        return

    result = st.session_state.calc_result
    color = "green" if result["status"] == "INNERHALB" else "red"
    print_status(f"{result['weekly']:.2f} → {result['status']}", color)

    if result["status"] == "ÜBERSCHRITTEN":
        with st.form("umrechnung_form", clear_on_submit=False):
            mm_pb = st.number_input("mm Pb:", min_value=0.0, step=0.1, key="mm_pb_input")
            convert = st.form_submit_button("Umrechnung")

        if convert:
            try:
                result = calculate(
                    raw_df,
                    df_data,
                    allowed=allowed,
                    mm_pb=mm_pb,
                    kv_override=kv_input,
                    out_pdf="Betriebsumfang_Auswertung.pdf",
                    write_pdf=True,
                )
                if result["conversion"]:
                    _, _, new_val, new_status, _ = result["conversion"]
                    color2 = "green" if new_status == "INNERHALB" else "red"
                    print_status(f"{new_val:.2f} → {new_status}", color2)

                    with open(result["out_pdf"], "rb") as f:
                        st.download_button(
                            "PDF herunterladen",
                            data=f.read(),
                            file_name="Betriebsumfang_Auswertung.pdf",
                            mime="application/pdf",
                            on_click=reset_for_next_entry,
                        )
                    print_download_message()
            except Exception as err:
                print_error(f"Umrechnung fehlgeschlagen: {err}")
    else:
        try:
            result = calculate(
                raw_df,
                df_data,
                allowed=allowed,
                mm_pb=None,
                kv_override=kv_input,
                out_pdf="Betriebsumfang_Auswertung.pdf",
                write_pdf=True,
            )
            with open(result["out_pdf"], "rb") as f:
                st.download_button(
                    "PDF herunterladen",
                    data=f.read(),
                    file_name="Betriebsumfang_Auswertung.pdf",
                    mime="application/pdf",
                    on_click=reset_for_next_entry,
                )
            print_download_message()
        except Exception as err:
            print_error(f"Fehler beim PDF-Erstellen: {err}")


if __name__ == "__main__":
    app()